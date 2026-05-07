import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
import tempfile
import datetime
from io import BytesIO

class ExcelExporter:
    def __init__(self, output_path="output.xlsx"):
        self.output_path = output_path

    def export_shots(self, shot_data):
        """
        shot_data: list of dicts from KitsuClient.get_all_shot_data
        """
        # Flatten tasks for processing (if needed) or keep it structured.
        # 엑셀에서는 한 샷 당 여러 줄 혹은 한 줄에 여러 컬럼으로 태스크를 표시할 수 있음.
        # 여기서는 한 줄에 여러 컬럼(태스크 타입별로) 표시하는 방식을 취함.
        
        # 태스크 타입 종류 파악
        task_types = set()
        for shot in shot_data:
            for task in shot["tasks"]:
                task_types.add(task["type"])
        
        sorted_task_types = sorted(list(task_types))
        
        # 데이터프레임 구성을 위한 리스트 생성
        rows = []
        for index, shot in enumerate(shot_data, start=1):
            row = {
                "No.": index,
                "Sequence": shot["sequence"],
                "Shot Name": shot["name"],
                "Description": shot["description"],
                "Frames": shot["nb_frames"],
                "Thumbnail": shot["thumbnail_url"] # URL 보관
            }
            # 태스크 정보 추가
            for tt in sorted_task_types:
                task_status = ""
                task_assignees = ""
                for t in shot["tasks"]:
                    if t["type"] == tt:
                        task_status = t["status"]
                        task_assignees = ", ".join(t.get("assignees", []))
                        break
                row[tt] = task_status
                row[tt + "__assignees"] = task_assignees
            rows.append(row)

        df = pd.DataFrame(rows)
        
        # Openpyxl을 사용하여 엑셀 작성 (이미지 삽입을 위해)
        wb = Workbook()
        ws = wb.active
        ws.title = "Shots"
        
        # 내용이 있는 부분 한페이지로 설정 (폭 기준)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # 헤더 작성: 각 태스크 타입 옆에 담당자 열 추가
        task_headers = []
        for tt in sorted_task_types:
            task_headers.append(tt)
            task_headers.append("Assignees")
        headers = ["No.", "Thumbnail", "Sequence", "Shot Name", "Description", "Frames"] + task_headers
        ws.append(headers)

        from openpyxl.styles import Alignment, Border, Side
        # 텍스트 중앙 정렬 및 자동 줄바꿈 적용
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 테두리 진하게 설정 (기존 thin_border 변수 재사용)
        thin_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        # 헤더 스타일 적용
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.alignment = center_alignment
            cell.border = thin_border

        # 행 데이터 및 이미지 처리
        temp_files = [] # 저장 후 삭제할 임시 파일 목록
        
        for i, row in enumerate(rows, start=2):
            # 텍스트 데이터 먼저 삽입
            c1 = ws.cell(row=i, column=1, value="=ROW()-1")
            c1.alignment = center_alignment
            c1.border = thin_border

            c2 = ws.cell(row=i, column=2) # 썸네일 들어갈 자리
            c2.alignment = center_alignment
            c2.border = thin_border

            c3 = ws.cell(row=i, column=3, value=row["Sequence"])
            c3.alignment = center_alignment
            c3.border = thin_border

            c4 = ws.cell(row=i, column=4, value=row["Shot Name"])
            c4.alignment = center_alignment
            c4.border = thin_border

            c5 = ws.cell(row=i, column=5, value=row["Description"])
            c5.alignment = center_alignment
            c5.border = thin_border

            c6 = ws.cell(row=i, column=6, value=row["Frames"])
            c6.alignment = center_alignment
            c6.border = thin_border

            for k, tt in enumerate(sorted_task_types):
                j = 7 + k * 2
                cell = ws.cell(row=i, column=j, value=row[tt])
                cell.alignment = center_alignment
                cell.border = thin_border
                cell_assignee = ws.cell(row=i, column=j + 1, value=row[tt + "__assignees"])
                cell_assignee.alignment = center_alignment
                cell_assignee.border = thin_border

            # 썸네일 이미지 처리 (preview_file_id 사용)
            preview_id = row["Thumbnail"]
            if preview_id:
                try:
                    import gazu
                    import urllib3
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

                    # Gazu 내부 클라이언트의 SSL 설정을 다시 한번 확정
                    try:
                        gazu.client.get_client().verify = False
                    except:
                        pass

                    # 임시 파일 생성
                    fd, tmp_path = tempfile.mkstemp(suffix=".png")
                    os.close(fd)
                    temp_files.append(tmp_path)
                    
                    try:
                        # Gazu 내장 함수로 고해상도 고정 이미지(Cover) 다운로드
                        # download_preview_file은 영상일 확률이 높으므로 커버를 우선 사용합니다.
                        print(f"DEBUG: Downloading high-quality cover for ID {preview_id}")
                        try:
                            # 1. 고화질 커버 시도 (영상의 대문 이미지 등)
                            gazu.files.download_preview_file_cover(preview_id, tmp_path)
                        except:
                            # 2. 실패 시 일반 썸네일 시도
                            print(f"DEBUG: Cover download failed, falling back to thumbnail for {row['Shot Name']}")
                            gazu.files.download_preview_file_thumbnail(preview_id, tmp_path)
                        
                        if os.path.exists(tmp_path) and os.path.getsize(tmp_path) > 0:
                            try:
                                img = Image(tmp_path)
                                # 이미지 크기 조정 (가로 210px, 세로는 비율에 맞게)
                                if img.width and img.height:
                                    aspect_ratio = img.height / img.width
                                    target_height = int(210 * aspect_ratio)
                                else:
                                    target_height = int(210 * 9 / 16) # 기본 16:9 비율
                                
                                img.width = 210
                                img.height = target_height
                                
                                # 5픽셀 정도의 여백 추가를 위한 Anchor 설정
                                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
                                from openpyxl.drawing.xdr import XDRPositiveSize2D
                                from openpyxl.utils.units import pixels_to_EMU
                                
                                # Column B는 index 1, Row i는 index i-1
                                marker = AnchorMarker(col=1, colOff=pixels_to_EMU(5), row=i-1, rowOff=pixels_to_EMU(5))
                                size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
                                img.anchor = OneCellAnchor(_from=marker, ext=size)
                                ws.add_image(img)
                                
                                # 엑셀 행 높이는 포인트 단위 (1픽셀 ≈ 0.75포인트), 위아래 약 5px씩 여백 추가 (10px * 0.75 = 7.5포인트 이상)
                                ws.row_dimensions[i].height = int(target_height * 0.75) + 10
                                print(f"DEBUG: Image embedded successfully for {row['Shot Name']}")
                            except Exception as img_err:
                                print(f"DEBUG: Openpyxl image loading error: {img_err}")
                        else:
                            print(f"DEBUG: No valid image file obtained for {row['Shot Name']}")
                    except Exception as download_err:
                        print(f"DEBUG: Overall download process error: {download_err}")
                            
                except Exception as e:
                    print(f"DEBUG: Error processing shot {row['Shot Name']}: {e}")

        # 열 너비 조절
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 27
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 10
        for k, tt in enumerate(sorted_task_types):
            from openpyxl.utils import get_column_letter
            j = 7 + k * 2
            ws.column_dimensions[get_column_letter(j)].width = 12     # 태스크 상태
            ws.column_dimensions[get_column_letter(j + 1)].width = 15  # 담당자

        # 모든 이미지 처리가 끝난 후 엑셀 저장 (이 시점까지 임시 파일이 유지되어야 함)
        print(f"DEBUG: Saving workbook to Downloads: {self.output_path}")
        wb.save(self.output_path)
        
        # 저장 완료 후 임시 파일 일괄 삭제
        for f_path in temp_files:
            try:
                if os.path.exists(f_path):
                    os.remove(f_path)
            except:
                pass
                
        print(f"DEBUG: Export completed. File saved at {self.output_path}")
        return self.output_path

    def export_delivery_list(self, delivery_data):
        """
        delivery_data: list of dicts with keys: name, thumbnail_url, version
        """
        from openpyxl.styles import Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery List"

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        today = datetime.date.today().strftime("%Y-%m-%d")
        headers = ["No.", "Thumbnail", "Shot Name", "Version", "Date"]
        ws.append(headers)

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.alignment = center_alignment
            cell.border = thin_border

        temp_files = []

        for i, item in enumerate(delivery_data, start=2):
            c1 = ws.cell(row=i, column=1, value="=ROW()-1")
            c1.alignment = center_alignment
            c1.border = thin_border

            c2 = ws.cell(row=i, column=2)
            c2.alignment = center_alignment
            c2.border = thin_border

            c3 = ws.cell(row=i, column=3, value=item["name"])
            c3.alignment = center_alignment
            c3.border = thin_border

            c4 = ws.cell(row=i, column=4, value=item["version"])
            c4.alignment = center_alignment
            c4.border = thin_border

            c5 = ws.cell(row=i, column=5, value=today)
            c5.alignment = center_alignment
            c5.border = thin_border

            preview_id = item["thumbnail_url"]
            if preview_id:
                try:
                    import gazu
                    import urllib3
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                    try:
                        gazu.client.get_client().verify = False
                    except:
                        pass

                    fd, tmp_path = tempfile.mkstemp(suffix=".png")
                    os.close(fd)
                    temp_files.append(tmp_path)

                    try:
                        try:
                            gazu.files.download_preview_file_cover(preview_id, tmp_path)
                        except:
                            gazu.files.download_preview_file_thumbnail(preview_id, tmp_path)

                        if os.path.exists(tmp_path) and os.path.getsize(tmp_path) > 0:
                            try:
                                img = Image(tmp_path)
                                if img.width and img.height:
                                    aspect_ratio = img.height / img.width
                                    target_height = int(210 * aspect_ratio)
                                else:
                                    target_height = int(210 * 9 / 16)

                                img.width = 210
                                img.height = target_height

                                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
                                from openpyxl.drawing.xdr import XDRPositiveSize2D
                                from openpyxl.utils.units import pixels_to_EMU

                                marker = AnchorMarker(col=1, colOff=pixels_to_EMU(5), row=i-1, rowOff=pixels_to_EMU(5))
                                size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
                                img.anchor = OneCellAnchor(_from=marker, ext=size)
                                ws.add_image(img)
                                ws.row_dimensions[i].height = int(target_height * 0.75) + 10
                            except Exception as img_err:
                                print(f"DEBUG: Image error for {item['name']}: {img_err}")
                    except Exception as download_err:
                        print(f"DEBUG: Download error for {item['name']}: {download_err}")
                except Exception as e:
                    print(f"DEBUG: Error processing {item['name']}: {e}")

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 27
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 15

        print(f"DEBUG: Saving delivery list to: {self.output_path}")
        wb.save(self.output_path)

        for f_path in temp_files:
            try:
                if os.path.exists(f_path):
                    os.remove(f_path)
            except:
                pass

        print(f"DEBUG: Delivery list export completed. File saved at {self.output_path}")
        return self.output_path
